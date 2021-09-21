# Jacob Holman general python functions

# trim_str, add_grey_line_TF, time_coverage, comma_unique, send_email, format_xl

# Added 2021-05-03
# start_end_tms, format_st_end_tms, tm_cov_pcnt, timed_input, Wait_Pixel

# Need to update time_coverage (and format)


# Date packages
from datetime import datetime
from datetime import timedelta
# Data handling packages
import pandas as pd
import numpy as np

import re, os
import pyperclip
# Packages for output to excel
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment, numbers

# Email stuff needed
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import os.path


def trim_str(x_str):
    # 2 or more spaces to 1
    # All leading or trailing spaces removed
    trm_str = re.sub(' +', ' ', x_str)
    trm_str = re.sub('^ +', '', trm_str)
    trm_str = re.sub(' +$', '', trm_str)
    return trm_str


def add_grey_line_TF(df, column):
    # Adds bool column alternating based on given column values
    # If value above in given column is the same then output will be the same bool value (T/F)
    # If not then output will be the opposite
    # First value is True
    
    # Tip: Data frame should be in disired order before using this function
    # Example usage: True false column can be used on excel for formatting
    df.reset_index(drop=True, inplace=True)
    grey_line = [True]
    for i in range(1, len(df[column])):
        if df[column][i] == df[column][i-1]:
            grey_line.append(grey_line[i-1])
        else:
            grey_line.append(not grey_line[i-1])
    
    df.insert(0, 'grey_line', grey_line)
    
    return df


def time_coverage(times, tm_delta_min = 60):
    # Given list of times
    # output will be string showing covered times e.g. "12:13-15:10, 17:01-17:34"
    # times in order are close enough defined by tm_delta_min (number of minutes)
    pc_tm = np.unique(times)
    pc_tm = sorted(pd.to_datetime(pc_tm))
    
    tm_prd = [pc_tm[0].strftime('%H:%M')]
    if len(pc_tm) > 1:
        for i in range(1, len(pc_tm)-1):
            if (pc_tm[i] - pc_tm[i-1]) > timedelta(minutes = tm_delta_min):
                if i >= 2:
                    if (pc_tm[i-1] - pc_tm[i-2]) > timedelta(minutes = tm_delta_min):
                        tm_prd.append(', ')
                        tm_prd.append(pc_tm[i].strftime('%H:%M'))
                    else:
                        tm_prd.append('-')
                        tm_prd.append(pc_tm[i-1].strftime('%H:%M'))
                        tm_prd.append(', ')
                        tm_prd.append(pc_tm[i].strftime('%H:%M'))
                else:
                    tm_prd.append(', ')
                    tm_prd.append(pc_tm[i].strftime('%H:%M'))
                        
        
        tm_prd.append('-')
        tm_prd.append(pc_tm[len(pc_tm)-1].strftime('%H:%M'))    
    
    return ''.join(tm_prd) 


def comma_unique(items):
    # Input is pandas series of strings
    # Output is comma separated unique values in order
    items_unq = np.unique(items)
    items_unq = [unq_itm for unq_itm in items_unq if str(unq_itm) not in ['nan', 'None', '']]
    return ', '.join(items_unq)


# Send email
def send_email(email_sender, email_recipients, email_subject, email_message, 
               attachment_location = ''):

    msg = MIMEMultipart()
    msg['From'] = email_sender
    msg['To'] = ','.join(email_recipients)
    msg['Subject'] = email_subject

    msg.attach(MIMEText(email_message, 'plain'))
    
    if attachment_location != '':
        filename = os.path.basename(attachment_location)
        attachment = open(attachment_location, "rb")
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', "attachment; filename= %s" % filename)
        msg.attach(part)
    
    try:
        server = smtplib.SMTP('erpsmtp')
        server.ehlo()
        text = msg.as_string()
        server.sendmail(email_sender, email_recipients, text)
        print('email sent')
        server.quit()
    except:
        input("SMPT server connection error " + chk_dt)
    return True


def format_xl(wb, sheet_nm, df, date_col, pcnt_col, cols_to_center, col_head_left, col_hide):
    ws = wb[sheet_nm]
    # Fill in Report sheet
    for r in dataframe_to_rows(df, index = False, header = True):
        ws.append(r)
    
    # Format headers
    for cell in ws[1]:
        cell.fill = PatternFill('solid', fgColor = '4F81BD')
        cell.font = Font(bold = True, size = 12, color = 'FFFFFF')
    
    # Format column as date
    for cell in ws[date_col]:
        cell.number_format = 'MM/DD/YYYY'
    
    # Format columns as percentage
    for col in pcnt_col:
        for cell in ws[col]:
            cell.number_format = numbers.FORMAT_PERCENTAGE
    
    # Center all values in specified columns
    for col in cols_to_center:
        for cell in ws[col]:
            cell.alignment = Alignment(horizontal = 'center')
    
    # Left align some column headers
    for col in col_head_left:
        ws[col][0].alignment = Alignment(horizontal = 'left')
    
    # Hide Columns
    for col in col_hide:
        ws.column_dimensions[col].hidden = True
    # Add Filter
    ws.auto_filter.ref = ws.dimensions
    
    return wb


# Format needed ts and bool 2 tuple
# gfi_data['test_tms1'] = list(zip(gfi_data['GFI_ts'], pd.Series(all_strt_ends)))
def start_end_tms(tm_st_end):
    # one parameter required for aggregation purposes
    logging.debug(tm_st_end)
    # tm_st_end is a list where each element is a time stamp and a bool tuple
    # The bool tuple indicates a start time or end time or neither
    
    # Make sure it is sorted or sort it
    st_end = pd.DataFrame(tm_st_end.tolist(), columns=['ts', 's_e']).sort_values(by = 'ts')
    st_end[['strt', 'end']] = pd.DataFrame(st_end['s_e'].tolist())
    
    if sum(st_end['end']) == 0:
        return []
    
    # Initialize list of time ranges
    in_tm = True
    s_e_times2 = []
    strt_tm = st_end['ts'][0]
    # Last Start, First End
    for index, row in st_end.iterrows():
        ts = row['ts']
        s_e = row['s_e']
        # Assert not both true
        assert not (s_e[0] and s_e[1]), 'Cannot have both start and end'
        # within time is true
        if s_e[1] and in_tm:
            in_tm = False
            s_e_times2.append((strt_tm, ts))
            logging.debug(str(s_e_times2))
        if s_e[0]:
            strt_tm = ts
            in_tm = True
    
    # This part is only needed when there is a known true start
    if in_tm and (list(tm_st_end)[-1][0] != strt_tm):
        s_e_times2.append((strt_tm, list(tm_st_end)[-1][0]))
        logging.debug(str(s_e_times2))
    
    return s_e_times2


def format_st_end_tms(tm_cov, frmt = '%H:%M'):
    frmt_rng_lst = []
    for rng in tm_cov:
        if rng[0] is None:
            frmt_rng_lst.append('-' + rng[1].strftime(frmt))
        elif len(rng) == 2:
            frmt_rng_lst.append(rng[0].strftime(frmt) + '-' + rng[1].strftime(frmt))
        else:
            raise Exception('Time coverage item is not a 2 tuple or datetime')
    
    return ', '.join(frmt_rng_lst)



# Time coverage percent
def tm_cov_pcnt(numer_time, denom_time):
    # Check denominator matches 
    ft_re = re.compile('[0-9]{2}:[0-9]{2}-[0-9]{2}:[0-9]{2}')
    if pd.isna(denom_time) or (not re.match(ft_re, denom_time)):
        return ''
    if pd.isna(numer_time) or numer_time == '':
        return 0
    
    strt_schd = datetime.strptime(denom_time[:5], '%H:%M')
    end_schd = datetime.strptime(denom_time[-5:], '%H:%M')
    sch_duration = end_schd - strt_schd
    sch_duration = sch_duration.total_seconds()
    
    # Avoid divide by 0 error
    sch_duration = .1 if sch_duration == 0 else sch_duration
    
    cov_secs = 0
    cov_list = numer_time.split(', ')
    for cov in cov_list:
        if '-' in cov:
            # Calculation fix.  -00:00 is actually midnight after but being read as midnight before
            cov = cov.replace('-00:00', '-23:59')
            strt_cov = datetime.strptime(cov[:5], '%H:%M')
            end_cov = datetime.strptime(cov[-5:], '%H:%M')
            # Make sure coverage time is before the end time of schedule
            if (strt_cov < end_schd) and (end_cov > strt_schd):
                if strt_cov < strt_schd:
                    strt_cov = strt_schd
                if end_cov > end_schd:
                    end_cov = end_schd
                
                cov_duration = (end_cov - strt_cov).total_seconds()
                cov_secs = cov_secs + cov_duration
    
    return cov_secs/sch_duration


def timed_input(helper_txt, time_limit):
    answer = None
    def check():
        time.sleep(time_limit)
        if answer != None:
            return
        print('too slow')
        os._exit(1)
    
    Thread(target = check).start()
    answer = input(helper_txt)
    return answer


def Wait_Pixel(pos, color, time_check = .1, timedout_CO = 10, time2skip = 31):
    # Pauses program until pixel shows in given location
    # input pos: 2 tuple int - pixel position
    # input color: 3 tuple int - RGB color
    # input time_check: float (.1) - time in seconds to wait to check again for pixel
    # input timedout_CO: float (10) - time in seconds until program raises exception
    # input time2skip: float (31) - time in seconds until waiting stops with no exception
    # packages: time, pyautogui as pag, logging
    time_elapse = 0
    while not pag.pixelMatchesColor(pos[0], pos[1], color):
        time.sleep(time_check)
        time_elapse += time_check
        if time_elapse >= timedout_CO:
            raise Exception('Waiting for Pixel timed out at ' + str(timedout_CO) + ' seconds')
        if time_elapse >= time2skip:
            logging.info('Wait Pixel has waited long enough and has stopeed waiting')
            break




